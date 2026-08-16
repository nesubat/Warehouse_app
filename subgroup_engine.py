import os
import json
import re
import xlwings as xw
from openpyxl.utils import get_column_letter
from core_math import generate_pack_signatures, update_metadata_for_subgroup, format_file1, format_file2, get_next_stage_filenames, close_if_open_elsewhere


class SubgroupValidationError(Exception):
    """Raised when the user-provided item row/numbers can't be safely resolved against the sheet."""
    pass


def _map_item_columns(row_data, tab_name, item_row):
    """Maps item numbers in a row to their column index.

    Aborts loudly (instead of silently overwriting/skipping) when the row has
    no item numbers at all, or when the same item number appears in more than
    one column (a typo, since item numbers must be unique per cell).
    """
    col_map = {}
    duplicates = {}

    for c_idx, val in enumerate(row_data):
        if val is None:
            continue
        try:
            item_num = int(float(val))
        except (ValueError, TypeError):
            continue

        col = c_idx + 1
        if item_num in col_map:
            duplicates.setdefault(item_num, [col_map[item_num]]).append(col)
        else:
            col_map[item_num] = col

    if not col_map:
        raise SubgroupValidationError(
            f"Tab '{tab_name}': no item numbers were found in row {item_row}. "
            f"Double-check that this is the correct Item Number Row."
        )

    if duplicates:
        details = "; ".join(
            f"item {num} appears in columns {', '.join(get_column_letter(c) for c in cols)}"
            for num, cols in duplicates.items()
        )
        raise SubgroupValidationError(
            f"Tab '{tab_name}': duplicate item numbers found in row {item_row} ({details}). "
            f"Item numbers must be unique per column — check for typos in the sheet and try again."
        )

    return col_map


def execute_subgroups(project_dir, metadata, subgroup_instructions):
    print("\n[DEBUG] =========================================")
    print("[DEBUG] STARTING SUB-GROUP ENGINE (XLWINGS)")
    print("[DEBUG] =========================================")
    
    file1_name = metadata.get("file1_name")
    if not file1_name:
        print("[ERROR] File 1 name missing from metadata.")
        return False
        
    job_id = metadata.get("job_id", "UNKNOWN")
    file_ext = os.path.splitext(file1_name)[1]
    
    
   # ---------------------------------------------------------
    # NEW: Generate dynamic Stage filenames
    # ---------------------------------------------------------
    # new_file1_name, new_json_name = get_next_stage_filenames(file1_name)
    base_file1 = re.sub(r"^Stage - \d+ - ", "", file1_name)
    name_only, ext = os.path.splitext(base_file1)
    
    # 2. Look inside the actual project folder to find the next available number
    stage_num = 1
    while True:
        test_file = f"Stage - {stage_num} - {base_file1}"
        if not os.path.exists(os.path.join(project_dir, test_file)):
            break
        stage_num += 1
        
    # 3. Assign the guaranteed unique names
    new_file1_name = f"Stage - {stage_num} - {base_file1}"
    new_json_name = f"Stage - {stage_num} - {name_only}.json"
    file1_path = os.path.join(project_dir, file1_name) # Open the old one
    new_file1_path = os.path.join(project_dir, new_file1_name) # Save as the new one
    new_json_path = os.path.join(project_dir, new_json_name)
    
    # Master Packing Sheet
    file2_name = f"Packing Sheet_{job_id}{file_ext}"
    file2_path = os.path.join(project_dir, file2_name)
    
    print(f"[DEBUG] Target File 1 (Source): {file1_name}")
    print(f"[DEBUG] New File 1 (Output): {new_file1_name}")
    print(f"[DEBUG] Target File 2 (Master Packing Sheet): {file2_name}")
    
    # Update metadata to reflect its new identity
    metadata["file1_name"] = new_file1_name
    
    app = xw.App(visible=False)
    app.display_alerts = False
    
    try:
        print("[DEBUG] Opening Excel files in background...")
        close_if_open_elsewhere(file1_path)
        close_if_open_elsewhere(file2_path)
        wb1 = app.books.open(file1_path)
        wb2 = app.books.open(file2_path)
        print("[DEBUG] Excel files opened successfully.")
        
        for tab_name, tab_data in subgroup_instructions.items():
            print(f"\n[DEBUG] --- Processing Tab: {tab_name} ---")
            if tab_name not in [s.name for s in wb1.sheets]:
                print(f"[DEBUG] Tab {tab_name} not found in File 1. Skipping.")
                continue
                
            ws1 = wb1.sheets[tab_name]
            meta_tab = metadata["tabs"][tab_name]
            item_row = tab_data["item_row"]
            pack_group_row = meta_tab["pack_group_row"]
            job_id_row = meta_tab["job_id_row"]
            last_row = meta_tab["last_row"]
            
            print(f"[DEBUG] Coordinates -> Item Row: {item_row}, Pack Row: {pack_group_row}, Job ID Row: {job_id_row}, Last Row: {last_row}")
            
            # =========================================================
            # STAGE 1: MEMORY EXTRACTION & MATH (didn't like the approach of mapping whole columns, so we will map item numbers to their respective columns by going through only pack selected.)
            # =========================================================
            print("[DEBUG] [STAGE 1] Extracting 2D Array and mapping item numbers...")
            raw_values = ws1.used_range.value

            if item_row - 1 >= len(raw_values):
                raise SubgroupValidationError(
                    f"Tab '{tab_name}': row {item_row} is out of range — this sheet only has "
                    f"{len(raw_values)} rows. Check the Item Number Row value."
                )

            item_row_data = raw_values[item_row - 1]
            item_col_map = _map_item_columns(item_row_data, tab_name, item_row)

            compiled_data = {}
            
            for pack_name, ranges in tab_data["packs"].items():
                print(f"[DEBUG] Computing Sub-groups for Parent Pack: {pack_name}")
                compiled_data[pack_name] = []
                for sg_index, (start_num, end_num) in enumerate(ranges):
                    start_col = item_col_map.get(int(start_num))
                    end_col = item_col_map.get(int(end_num))
                    start_col_letter = get_column_letter(start_col) if start_col else "N/A"
                    end_col_letter = get_column_letter(end_col) if end_col else "N/A"   
                    
                    if not start_col or not end_col:
                        missing = [str(n) for n, c in ((start_num, start_col), (end_num, end_col)) if not c]
                        raise SubgroupValidationError(
                            f"Tab '{tab_name}', Pack '{pack_name}': item number(s) {', '.join(missing)} "
                            f"not found in row {item_row}. Check for typos in the Start/End Item # "
                            f"fields or the Item Number Row."
                        )

                    print(f"[DEBUG]   -> Sub-group {start_num}-{end_num} mapped to Columns {start_col_letter} through {end_col_letter}")
                    store_rows = range(pack_group_row + 1, last_row + 1)
                    
                    row_sigs, unique_sigs, sig_to_letter, summary_counts, ordered_codes = generate_pack_signatures(
                        raw_values, store_rows, start_col, end_col
                    )
                    print(f"[DEBUG]   -> Math complete: Found {len(unique_sigs)} unique signatures.")
                    
                    compiled_data[pack_name].append({
                        "header_str": f"{start_num}-{end_num}",
                        "start_num": start_num,
                        "end_num": end_num,
                        "start_col": start_col,
                        "end_col": end_col,
                        "unique_sigs": unique_sigs,
                        "summary_counts": summary_counts, # FIX: Matches Stage 3
                        "sig_to_letter": sig_to_letter,   # FIX: Matches Stage 3
                        "ordered_codes": ordered_codes,
                        "row_sigs": row_sigs
                        })

            # =========================================================
            # STAGE 2: RIGHT-TO-LEFT INSERTION (FILE 1)
            # =========================================================
            print("\n[DEBUG] [STAGE 2] Sorting packs for Right-To-Left insertion...")
            packs_to_process = []
            for p_name in compiled_data.keys():
                c_col = meta_tab["packs"][p_name]["code_col"]
                packs_to_process.append((c_col, p_name))
                
            packs_to_process.sort(key=lambda x: x[0], reverse=True)
            print(f"[DEBUG] Insertion order: {[p[1] for p in packs_to_process]}")
            
            for code_col_orig, p_name in packs_to_process:
                subgroups = compiled_data[p_name]
                current_pstart = meta_tab["packs"][p_name]["code_col"]
                insert_base = current_pstart + 1
                original_color = ws1.range((pack_group_row, insert_base - 1)).color
                current_pstart_col = get_column_letter(current_pstart)
                current_pend = meta_tab["packs"][p_name]["current_end"]
                current_pend_col = get_column_letter(current_pend)
                
                print(f"[DEBUG] Modifying File 1 for Pack: {p_name} | Base Insertion Col: {insert_base}")
                # Unmerge the cell at the pack group row for the original code column to prepare for new insertions
                ws1.range(f"{current_pstart_col}{pack_group_row}:{current_pend_col}{pack_group_row}").unmerge()
                # Insert new columns and batch write sigs for each sub-group
                
                for sg_index, sg in enumerate(subgroups):
                    current_insert_idx = insert_base + sg_index
                    col_letter = get_column_letter(current_insert_idx)
                    
                    print(f"[DEBUG]   -> Inserting column {col_letter} for Sub-group {sg['header_str']}")
                    ws1.range(f"{col_letter}:{col_letter}").insert('right')
                    ws1.range(f"{col_letter}{job_id_row}").value = f"Code for {sg['header_str']}"
                    
                    
                    format_file1(ws1, col_letter, current_pstart, sg['end_col'], pack_group_row, job_id_row, last_row)
                    # pulling the row signatures and writing them in batch to the newly inserted column
                    row_sigs= sg['row_sigs']
                    sig_to_letter= sg['sig_to_letter']
                    batch_data=[]
                    for r, sig in row_sigs:
                        if sig is not None:
                            batch_data.append([sig_to_letter[sig]])
                        else:
                            batch_data.append([None])
                    # Write the batch data to the newly inserted column
                    if batch_data:
                        start_row = row_sigs[0][0]  # First row number
                        ws1.range(f"{col_letter}{start_row}").value = batch_data
                        print(f"[DEBUG]   -> Wrote {len(batch_data)} rows of signature codes to column {col_letter} starting at row {start_row}.")

                    metadata = update_metadata_for_subgroup(
                        metadata, tab_name, p_name, sg['header_str'], current_insert_idx, sg['start_col'], sg['end_col']
                    )
                
                new_pend_col=get_column_letter(current_pend + len(subgroups))
                new_pack_range=ws1.range(f"{current_pstart_col}{pack_group_row}:{new_pend_col}{pack_group_row}")
                new_pack_range.merge()
                new_pack_range.color=original_color

            # =========================================================
            # STAGE 3: SIDE-BY-SIDE MATRIX LAYOUT (FILE 2)
            # =========================================================
            print("\n[DEBUG] [STAGE 3] Generating Side-by-Side Matrix Layout in File 2...")
            file2_matrices = []
            for p_name, ranges in tab_data["packs"].items():
                if p_name in compiled_data:
                    for sg_mat in compiled_data[p_name]:
                        file2_matrices.append((p_name, sg_mat))
            job_id_row =meta_tab["job_id_row"]

            if file2_matrices:
                sg_tab_name = f"{tab_name[:26]} - SG"
                ws2_orig = wb2.sheets[tab_name]
        
                    # Check if the sub-group tab already exists
                if sg_tab_name in [s.name for s in wb2.sheets]:
                    print(f"[DEBUG] Reusing existing tab: {sg_tab_name}")
                    ws2_sg = wb2.sheets[sg_tab_name]
                    last_col = ws2_sg.range(job_id_row, ws2_sg.cells.last_cell.column).end('left').column
                    
                    offset_col = last_col + 2

                            
                else:
                    # Tab doesn't exist yet, so create it
                    print(f"[DEBUG] Created new tab: {sg_tab_name}")
                    ws2_sg = wb2.sheets.add(name=sg_tab_name, after=ws2_orig)
                    offset_col = 1

                # --- ADD THIS: Re-map item columns after Stage 2 insertions ---
                print("[DEBUG] Re-mapping item columns in File 1 after insertions...")
                # Pull the freshly updated item row from File 1
                updated_item_row_data = ws1.range(f"{item_row}:{item_row}").value
                updated_item_col_map = _map_item_columns(updated_item_row_data, tab_name, item_row)
                # --------------------------------------------------------------

                for p_name, sg_mat in file2_matrices:
                    # Dynamically grab the newly shifted columns
                    s_col = updated_item_col_map.get(int(sg_mat['start_num']))
                    e_col = updated_item_col_map.get(int(sg_mat['end_num']))
                    
                    if not s_col or not e_col:
                        raise SubgroupValidationError(
                            f"Tab '{tab_name}': could not find shifted columns for sub-group "
                            f"'{sg_mat['header_str']}' after inserting columns. This points to an "
                            f"internal mapping error — please report this."
                        )

                    col_span = e_col - s_col
                    print(f"[DEBUG] Building Matrix for {p_name} ({sg_mat['header_str']}) at offset column {offset_col}")

                    # 1. COPY EXACT METADATA SNIPPET
                    print(f"[DEBUG]   -> Copying metadata snippet from File 1 (Cols {get_column_letter(s_col)} to {get_column_letter(e_col)})")
                    
                    
                    # IF YOU ONLY WANT TO COPY THE SINGLE PACK GROUP ROW:
                    snippet_range = ws1.range((1, s_col), (pack_group_row-1, e_col))
                    snippet_range.copy(ws2_sg.range((1, offset_col)))
                    merge_range = ws2_sg.range((pack_group_row, offset_col), (pack_group_row, offset_col + col_span))

                    # 2. Write the header for the sub-group matrix
                    ws2_sg.range((job_id_row, offset_col + col_span + 1)).value = "Count"
                    ws2_sg.range((job_id_row, offset_col + col_span + 2)).value = f"Code for {sg_mat['header_str']}"
                    ws2_sg.range((job_id_row, offset_col + col_span + 1), (job_id_row, offset_col + col_span + 2)).font.size =ws2_sg.range((job_id_row, offset_col)).font.size

                    # Merge the cells
                    merge_range.merge()

                    # Write the text (using the header_str which already contains "start_num-end_num")
                    merge_range.value = f"Code for {sg_mat['header_str']}"

                    # (Optional) Center the text horizontally so it looks clean across the merged cells
                    merge_range.api.HorizontalAlignment = -4108

                    original_font_size = ws1.range((pack_group_row, s_col)).font.size
                    original_font_color = ws1.range((pack_group_row, s_col)).font.color
                    original_color = ws1.range((pack_group_row, s_col)).color
                    merge_range.color = original_color
                    merge_range.font.size = original_font_size
                    # Carry over row heights
                    for r in range(1, pack_group_row + 1):
                        ws2_sg.range((r, 1)).row_height = ws1.range((r, 1)).row_height

                    # Carry over column widths
                    for i in range(col_span + 1):
                        ws2_sg.range((1, offset_col + i)).column_width = ws1.range((1, s_col + i)).column_width
                    

                    
                    

                    # 3. BUILD MATRIX DATA
                    print("[DEBUG]   -> Compiling Matrix array...")
                    write_data = []
                    num_items = len(sg_mat["unique_sigs"][0])
                    totals = [0] * (num_items + 1)

                    for sig in sg_mat["unique_sigs"]:
                        row_data = [val if val != 0 else None for val in sig]
                        # THE FIX: Updated dictionary keys to match Stage 1
                        row_data.append(sg_mat["summary_counts"][sig])
                        row_data.append(sg_mat["sig_to_letter"][sig])
                        write_data.append(row_data)

                        for i, val in enumerate(sig):
                            if val:
                                try:
                                    # THE FIX: Updated dictionary key here too
                                    totals[i] += (float(val) * sg_mat["summary_counts"][sig])
                                except (ValueError, TypeError):
                                    pass
                        # THE FIX: Updated dictionary key here too
                        totals[-1] += sg_mat["summary_counts"][sig]

                    total_row = [t if t != 0 else None for t in totals[:-1]]
                    total_row.append(totals[-1])
                    total_row.append("Total")
                    write_data.append(total_row)

                    # 4. INJECT AND FORMAT
                    print("[DEBUG]   -> Injecting Matrix array and applying formatting...")
                    format_file2(ws2_sg, pack_group_row, offset_col, offset_col + col_span, write_data)
                    ws2_sg.range((job_id_row, offset_col + col_span + 1), (job_id_row, offset_col + col_span + 2)).api.EntireColumn.AutoFit()

                    last_col = offset_col + col_span + 2
                    wall_col = last_col + 1
                    ws2_sg.range(f"{get_column_letter(wall_col)}:{get_column_letter(wall_col)}").column_width = 1
                    ws2_sg.range(f"{get_column_letter(wall_col)}:{get_column_letter(wall_col)}").color=(0,0,0)
                    offset_col = last_col + 2  # Move to the next starting column for the next matrix
        # =========================================================  

        print("\n[DEBUG] Saving and closing workbooks...")
        wb1.save(new_file1_path)
        wb1.close()
        wb2.save()
        wb2.close()
        print("[DEBUG] Sub-group execution completed successfully!")
        
    except Exception as e:
        print(f"[FATAL ERROR] Sub-group execution failed: {e}")
        import traceback
        traceback.print_exc()
        raise e
    finally:
        try:
            app.quit()
            app.kill()
        except:
            pass

    # Save Metadata Tracker
    meta_path = os.path.join(project_dir, "project_metadata.json")
    with open(new_json_path, 'w') as f:
        json.dump(metadata, f, indent=4)
        
    return True