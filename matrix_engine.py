import os
import shutil
import re
import openpyxl
import xlwings as xw
import pandas as pd
import collections
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
import json
from core_math import clean_file_name, generate_pack_signatures, format_file1, format_file2, build_initial_metadata, update_metadata_for_subgroup


def scan_excel_tabs(file_path):
    excel_file = pd.ExcelFile(file_path)
    excel_file.close()
    return excel_file.sheet_names

def generate_tab_map(file_path, sheet_name, start_cell, job_id_cell, store_col):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        sheet = wb[sheet_name]
    
        start_coords = coordinate_from_string(start_cell) 
        stock_start_col = column_index_from_string(start_coords[0])
        pack_group_row = start_coords[1]
        
        job_coords = coordinate_from_string(job_id_cell)
        pack_start_col = column_index_from_string(job_coords[0])
        job_id_row = job_coords[1]
        
        raw_job_id = sheet[job_id_cell].value
        store_col_idx = column_index_from_string(store_col)
        
        true_max_col = 1
        for col in range(sheet.max_column, 0, -1):
            cell_val = sheet.cell(row=job_id_row, column=col).value
            if cell_val is not None and str(cell_val).strip() != "":
                true_max_col = col
                break
                
        max_col = true_max_col
        last_col_letter = get_column_letter(max_col)
            
        last_row = sheet.max_row
        while last_row > 0 and sheet.cell(row=last_row, column=store_col_idx).value is None:
            last_row -= 1
        if "total" or "grand total" in str(sheet.cell(row=last_row, column=store_col_idx).value).strip().lower():
            last_row -= 1
            
        total_stores = last_row - max(pack_group_row, job_id_row)

        # --- DUPLICATE STORE SCANNER ---
        store_names = []
        start_store_row = max(pack_group_row, job_id_row) + 1
        
        for r in range(start_store_row, last_row + 1):
            val = sheet.cell(row=r, column=store_col_idx).value
            if val is not None and str(val).strip() != "":
                store_names.append(str(val).strip())
        
        duplicate_warning = None
        if len(store_names) != len(set(store_names)):
            dupes = [item for item, count in collections.Counter(store_names).items() if count > 1]
            duplicate_warning = f"⚠️ Duplicate Store Names Detected: {', '.join(dupes[:3])}{'...' if len(dupes)>3 else ''}. This will cause grouping collisions while shuffling the labels!"
        
        last_stock_col = stock_start_col
        while sheet.cell(row=pack_group_row, column=last_stock_col + 1).value is not None:
            last_stock_col += 1
            
        stock_start_ltr = get_column_letter(stock_start_col)
        stock_end_ltr = get_column_letter(last_stock_col)
            
        last_pack_col = max_col
        while last_pack_col > pack_start_col and sheet.cell(row=job_id_row, column=last_pack_col).value is None:
            last_pack_col -= 1
            
        packages = []
        pack_ranges = [] 
        current_col = pack_start_col  
        
        while current_col <= last_pack_col:
            is_merged = False
            
            for merged_range in sheet.merged_cells.ranges:
                if (current_col >= merged_range.min_col and current_col <= merged_range.max_col and 
                    pack_group_row >= merged_range.min_row and pack_group_row <= merged_range.max_row):
                    
                    raw_name = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                    pack_name = str(raw_name).strip() if raw_name else f"Pack_{current_col}"
                    
                    start_letter = get_column_letter(merged_range.min_col)
                    end_letter = get_column_letter(merged_range.max_col)
                    
                    packages.append({
                        "name": pack_name, 
                        "label": f"{pack_name} spans from {start_letter} to {end_letter}"
                    })
                    pack_ranges.append({
                        "name": pack_name, 
                        "start": merged_range.min_col, 
                        "end": merged_range.max_col
                    })
                    
                    current_col = merged_range.max_col + 1 
                    is_merged = True
                    break
                    
            if not is_merged:
                raw_name = sheet.cell(row=pack_group_row, column=current_col).value
                if raw_name is not None and str(raw_name).strip() != "":
                    pack_name = str(raw_name).strip()
                    start_letter = get_column_letter(current_col)
                    
                    packages.append({
                        "name": pack_name,
                        "label": f"{pack_name} is in column {start_letter}"
                    })
                    pack_ranges.append({
                        "name": pack_name, 
                        "start": current_col, 
                        "end": current_col
                    })
                    
                current_col += 1
                
        backend_data = {
            "job_id_row": job_id_row,
            "pack_group_row": pack_group_row,
            "last_row": last_row,
            "pack_ranges": pack_ranges,
            "store_col_idx": store_col_idx,
            "stock_start_col": stock_start_col,
            "last_stock_col": last_stock_col,
            "raw_job_id": raw_job_id
        }
                
    finally:
        wb.close()
            
    return {
        "sheet_name": sheet_name,
        "stocks_map": f"Packaging stocks are in {stock_start_ltr} to {stock_end_ltr}.",
        "packages_map": packages,
        "last_col": last_col_letter,
        "last_row": last_row,
        "total_stores": total_stores,
        "duplicate_warning": duplicate_warning,
        "backend_data": backend_data
    }

# =====================================================================
# THE CORE GENERATOR ENGINE (PHASE 1, 2, and 3)
# =====================================================================

def generate_all_outputs(file_path, original_filename, selected_tabs, user_inputs, tab_data_memory, project_dir):
    file_ext = os.path.splitext(file_path)[1]
    
    first_tab = selected_tabs[0]
    raw_job_id = tab_data_memory[first_tab]["raw_job_id"]
    cleaned_job_id = clean_file_name(raw_job_id)
    
    # ---------------------------------------------------------
    # MASTER SWITCH: Check if any tab has any packs selected
    # ---------------------------------------------------------
    any_packs_selected = False
    for tab in selected_tabs:
        inputs = user_inputs.get(tab, {})
        if inputs.get("selected_packs"):
            any_packs_selected = True
            break

    file1_name = f"Final {os.path.splitext(original_filename)[0]}{file_ext}" if any_packs_selected else None
    file2_name = f"Packing Sheet_{cleaned_job_id}{file_ext}"
    file3_name = f"Signature links_{cleaned_job_id}{file_ext}" if any_packs_selected else None
    
    file1_path = os.path.join(project_dir, file1_name) if any_packs_selected else None
    file2_path = os.path.join(project_dir, file2_name)
    file3_path = os.path.join(project_dir, file3_name) if any_packs_selected else None

    # --- NEW: INITIALIZE METADATA TRACKER ---
    project_metadata = {
        "job_id": cleaned_job_id,
        "file1_name": file1_name, 
        "tabs": {}
    }
    
    if any_packs_selected:
        shutil.copy(file_path, file1_path)
    shutil.copy(file_path, file2_path)
    
    tab_summaries = {}
    app = xw.App(visible=False)
    app.display_alerts = False
    
    try:
        # =====================================================
        # PHASE 1 & 2: GENERATE FILE 1 & 2 SIMULTANEOUSLY 
        # =====================================================
        if any_packs_selected:
            wb1_xw = app.books.open(file1_path)
            
        wb2_xw = app.books.open(file2_path)
        for sheet in list(wb2_xw.sheets):
            if sheet.name not in selected_tabs:
                sheet.delete()
        master_stock_data = []
        
        for tab_name in selected_tabs:
            sheet2 = wb2_xw.sheets[tab_name]
            tab_info = tab_data_memory[tab_name]
            raw_values = sheet2.used_range.value 
            
            tab_summaries[tab_name] = []
            inputs = user_inputs.get(tab_name, {"selected_packs": []})
            selected_list = [p.strip() for p in inputs.get("selected_packs", [])]

            # --- NEW: BUILD METADATA USING DECOUPLED ENGINE ---
            project_metadata["tabs"][tab_name] = build_initial_metadata(tab_info, inputs, selected_list, any_packs_selected)
            
            if any_packs_selected:
                sheet1_xw = wb1_xw.sheets[tab_name]
            
            # --- CALCULATE SIGNATURES & INJECT INTO FILE 1 (If Selected) ---
            for pack in reversed(tab_info["pack_ranges"]):
                p_name = pack["name"]
                p_start = pack["start"]
                p_end = pack["end"]
                job_id_row = tab_info["job_id_row"]

                store_rows = range(tab_info["pack_group_row"] + 1, tab_info["last_row"] + 1)
                
               # CALL THE NEW DECOUPLED ENGINE
                row_signatures, unique_sigs, sig_to_letter, summary_counts, ordered_codes = generate_pack_signatures(raw_values, store_rows, p_start, p_end)
                print(f"Processed pack '{p_name}' in tab '{tab_name}': {len(unique_sigs)} unique signatures found.")
                    
                is_pack_selected = p_name.strip() in selected_list
                
                if any_packs_selected and is_pack_selected:
                    col_letter_start = get_column_letter(p_start)
                    col_letter_end = get_column_letter(p_end)
                    # unmerging the pack group row before inserting the new column to avoid merge conflicts
                    sheet1_xw.range(f"{col_letter_start}{tab_info['pack_group_row']}:{col_letter_end}{tab_info['pack_group_row']}").unmerge()
                    sheet1_xw.range(f"{col_letter_start}:{col_letter_start}").insert('right')
                    
                    pack_group_row = tab_info["pack_group_row"]
                    pack_name_cell = sheet1_xw.range((pack_group_row, p_start+1))
                    original_color = pack_name_cell.color
                    sheet1_xw.range(f"{col_letter_start}{job_id_row}").value = f"Code for Pack {p_name}"
                    

                    
                    # --- CALL DECOUPLED FORMATTING FOR FILE 1 ---
                    format_file1(
                        sheet1_xw, col_letter_start, p_start, p_end, 
                        pack_group_row, tab_info['job_id_row'], tab_info['last_row']
                    )
                    print(f"Inserted and formatted signature column for pack '{p_name}' in tab '{tab_name}'.")

                    # --- THE BATCH WRITE FIX ---
                    # Bundle all the letters into a 2D array (a vertical column list)
                    batch_data = []
                    for r, sig in row_signatures:
                        if sig is not None:
                            batch_data.append([sig_to_letter[sig]])
                        else:
                            batch_data.append([None])
                            
                    # Paste the entire column into Excel in exactly ONE command!
                    if batch_data:
                        start_r = row_signatures[0][0] # Grab the very first row number
                        sheet1_xw.range(f"{col_letter_start}{start_r}").value = batch_data
                        print(f"Batch wrote {len(batch_data)} signature codes for pack '{p_name}' in tab '{tab_name}'.")
                    # =========================================================
                    # NEW: RE-MERGE MAIN PACK HEADER AFTER BATCH WRITE
                    # =========================================================
                    col_letter_new_end = get_column_letter(p_end + 1)
                    new_pack_range = sheet1_xw.range(f"{col_letter_start}{pack_group_row}:{col_letter_new_end}{pack_group_row}")
                    
                    new_pack_range.merge()
                    if original_color:
                        new_pack_range.color = original_color
                    new_pack_range.api.HorizontalAlignment = -4108
                    new_pack_range.api.VerticalAlignment = -4108
                
                        
                            
                tab_summaries[tab_name].append({
                    "name": p_name,
                    "start_idx": p_start,
                    "end_idx": p_end,
                    "unique_sigs": unique_sigs,
                    "counts": summary_counts,
                    "letters": sig_to_letter,
                    "is_selected": is_pack_selected,
                    "ordered_codes": ordered_codes
                })

            # --- BUILD PACKING SHEET SUMMARY (FILE 2) ---
            pack_group_row = tab_info["pack_group_row"]
            job_id_row = tab_info["job_id_row"]
            
            stock_start_col = tab_info["stock_start_col"]
            for idx, raw_pack in enumerate(tab_info["pack_ranges"]):
                stock_col = stock_start_col + idx
                stock_values = sheet2.range((pack_group_row + 1, stock_col), (tab_info["last_row"], stock_col)).value
                if not isinstance(stock_values, list): stock_values = [stock_values]
                
                stock_counts = {}
                for val in stock_values:
                    if val: 
                        v_str = str(val).strip()
                        stock_counts[v_str] = stock_counts.get(v_str, 0) + 1
                        
                master_stock_data.append({
                    "header": f"{tab_name}_{raw_pack['name']}",
                    "counts": stock_counts
                })

            start_del = pack_group_row + 1
            sheet2.range(f"{start_del}:1048576").api.EntireRow.Delete()
            
            summaries = tab_summaries.get(tab_name, [])
            for p_sum in summaries:
                p_name = p_sum["name"]
                
                raw_pack = next(p for p in tab_info["pack_ranges"] if p["name"] == p_name)
                p_start = raw_pack["start"]
                p_end = raw_pack["end"]
                
                col_letter_1 = get_column_letter(p_end + 1)
                col_letter_2 = get_column_letter(p_end + 2)
                
                sheet2.range(f"{col_letter_1}:{col_letter_2}").insert('right')
                sheet2.range(f"{col_letter_1}:{col_letter_2}").color = None
                    
                sheet2.range((job_id_row, p_end + 1)).value = "Count"
                sheet2.range((job_id_row, p_end + 2)).value = f"Code for {p_name}"
                
                # --- 1. AUTOFIT THE NEW COLUMNS ---
                sheet2.range(f"{col_letter_1}:{col_letter_2}").api.EntireColumn.AutoFit()
                
                
                write_data = []
                num_items = len(p_sum["unique_sigs"][0])
                totals = [0] * (num_items + 1) 
                
                for sig in p_sum["unique_sigs"]:
                    row_data = [val if val != 0 else None for val in sig]
                    row_data.append(p_sum["counts"][sig])   
                    row_data.append(p_sum["letters"][sig])  
                    write_data.append(row_data)
                    
                    for i, val in enumerate(sig):
                        if val: 
                            try:
                                numeric_val = float(val)
                                totals[i] += (val * p_sum["counts"][sig])
                            except (ValueError, TypeError):
                                pass  
                    totals[-1] += p_sum["counts"][sig]
                
                total_row = [t if t != 0 else None for t in totals[:-1]]
                total_row.append(totals[-1])
                total_row.append("Total")
                write_data.append(total_row)
                    
                # --- CALL DECOUPLED FORMATTING FOR FILE 2 ---
                format_file2(sheet2, pack_group_row, p_start, p_end, write_data) 
                            
            end_del_col_letter = get_column_letter(tab_info["last_stock_col"] + 1)
            sheet2.range(f"A:{end_del_col_letter}").api.EntireColumn.Delete()

            # Preparing the stock summary for the Packaging Stocks sheet

        stock_sheet = wb2_xw.sheets.add(name="Packaging Stocks", after=wb2_xw.sheets[-1])
        out_row = 1

        for stock_info in master_stock_data:
            header_row = out_row
            stock_sheet.range(f"A{header_row}:B{header_row}").merge()
            stock_sheet.range(f"A{header_row}").value = stock_info["header"]
            stock_sheet.range(f"A{header_row}").color = (217, 234, 211)

            row_idx = header_row + 1
            for stock_name, count in stock_info["counts"].items():
                stock_sheet.range((row_idx, 1)).value = stock_name
                stock_sheet.range((row_idx, 2)).value = count
                row_idx += 1

            wall_row = row_idx
            stock_sheet.range(f"{wall_row}:{wall_row}").color = (0, 0, 0)
            stock_sheet.range(f"{wall_row}:{wall_row}").row_height = 6

            out_row = wall_row + 1

        

        # Uniform font size + full grid borders across everything just written
        last_used_row = out_row - 2  # step back past the final trailing divider row
        if last_used_row >= 1:
            used_range = stock_sheet.range((1, 1), (last_used_row, 2))
            used_range.font.size = 20
            for border_id in [7, 8, 9, 10, 11, 12]:
                used_range.api.Borders(border_id).LineStyle = 1
                used_range.api.Borders(border_id).Weight = 2
                
        stock_sheet.range("A:B").api.EntireColumn.AutoFit()
                    
        if any_packs_selected:
            wb1_xw.save()
            wb1_xw.close()
            
        wb2_xw.save()
        wb2_xw.close()
    finally: 
        try:
            app.quit()
            app.kill()
        except:
            pass

    # =====================================================
    # PHASE 3: GENERATE FILE 3 (Pandas Rebuild Architecture)
    # =====================================================
    if any_packs_selected:
        wb_raw = openpyxl.load_workbook(file_path, data_only=True)
        
        with pd.ExcelWriter(file3_path, engine='xlsxwriter') as writer:
            for tab_name in selected_tabs:
                tab_info = tab_data_memory[tab_name]
                sheet_raw = wb_raw[tab_name]
                
                # Check if this specific tab has any selected packs
                has_codes = any(p_sum["is_selected"] for p_sum in tab_summaries.get(tab_name, []))
                if not has_codes:
                    continue # Skip writing this tab entirely if it has no codes
                
                store_col_idx = tab_info["store_col_idx"]
                start_r = tab_info["pack_group_row"] + 1
                end_r = tab_info["last_row"]
                
                # Extract Stores
                store_names = []
                for r in range(start_r, end_r + 1):
                    val = sheet_raw.cell(row=r, column=store_col_idx).value
                    store_names.append(str(val) if val is not None else "")
                
                df_dict = {"Store Name": store_names}
                
                # Extract Ordered Codes
                for p_sum in reversed(tab_summaries.get(tab_name, [])):
                    if p_sum["is_selected"]:
                        # Reverse list structure to write cleanly
                        p_name = p_sum["name"]
                        ordered_codes = p_sum["ordered_codes"]
                        # Pad or trim if lengths mismatch (failsafe)
                        if len(ordered_codes) < len(store_names):
                            ordered_codes.extend([""] * (len(store_names) - len(ordered_codes)))
                        elif len(ordered_codes) > len(store_names):
                            ordered_codes = ordered_codes[:len(store_names)]
                            
                        df_dict[f"Code for {p_name}"] = ordered_codes
                
                # Write to Excel
                df_out = pd.DataFrame(df_dict)
                df_out.to_excel(writer, sheet_name=tab_name, index=False)
                
                # Apply Formatting via xlsxwriter engine
                worksheet = writer.sheets[tab_name]
                workbook = writer.book
                format16 = workbook.add_format({'font_size': 16, 'align': 'center'})
                
                # Set Autofit Width to 25 and force Font Size 16
                worksheet.set_column(0, len(df_dict) - 1, 25, format16) 
                
        wb_raw.close()
    # --- NEW: SAVE METADATA JSON TO PROJECT FOLDER ---
    
    base_file1_name = os.path.splitext(file1_name)[0]
    metadata_path = os.path.join(project_dir, f"{base_file1_name}.json")
    with open(metadata_path, 'w') as f:
        json.dump(project_metadata, f, indent=4)
    
    return file1_name, file2_name, file3_name